import streamlit as st
import pandas as pd
import numpy as np
import openpyxl as xl
import matplotlib.pyplot as plt
from generate_sql_from_template import generate_sql
from sql_execution import execute_sql, get_execution_dates, generate_summ_report, generate_det_report
from database_connection import snowflake_acct_connect, snowflake_ws_connect
from config import get_ini_sections

# Create snowflake objects
def sql_generator_streamlit(username):
    uploaded_file = None
    # session objects to store snow flake connection
    if 'acc_con' not in st.session_state:
        st.session_state.acc_con = None

    if 'ws_con' not in st.session_state:
        st.session_state.ws_con = None
    acct_list = get_ini_sections()  # get list of accounts from config.ini file
    acct_option = st.selectbox('Account', options=acct_list)
    btn_connect = st.button("Connect")
    if btn_connect:
        if acct_option == '<select account>':
            st.warning("Please select account")
        else:
            st.session_state.acc_con = snowflake_acct_connect(acct_option) # get connection object without warehouse
            st.session_state.ws_con = snowflake_ws_connect(acct_option) # get connection object without warehouse
            if st.session_state.acc_con is not None and st.session_state.ws_con is not None:
                st.success("Connected successfully")
            else:
                st.success("Connection failed")
    # template upload and processing section

    uploaded_file = st.file_uploader("Upload template", type=['xlsx'])

    btn_submit = st.button('Submit')
    if btn_submit:
        if uploaded_file is None:
            st.info("Please upload template")
        elif st.session_state.acc_con is None:
            st.info("Please connect to an account using Admin tab")
        else:
            wb = xl.load_workbook(uploaded_file)
            sheets = wb.sheetnames
            res = len(wb.sheetnames)
            final_sql_list = []
            for i in range(res):
                if sheets[i] != "SHARED_DB" and sheets[i] != "STAGE" and sheets[i] != "Lookup":
                    with st.spinner("Generating SQL commands. Please wait..."):
                        df = pd.read_excel(
                            uploaded_file,
                            sheets[i]
                        )
                        df.replace(np.NaN, '', inplace=True)
                        sql_list = generate_sql(df, sheets[i]) # get sql commands from template excel
                        final_sql_list.extend(sql_list)

            with st.spinner("Executing SQL commands. Please wait..."):
                execute_sql(final_sql_list,"nxljbzs-qg26858",username,st.session_state.acc_con, st.session_state.ws_con) # execute sql commandes "nxljbzs-qg26858" need to be replaced with acct_option
                st.success("Execution completed")

#Ad-Hoc Section
def adhoc(username):
    acct_list = get_ini_sections()  # get list of accounts from config.ini file
    acct_option = st.selectbox('Account', options=acct_list)
    btn_connect = st.button("Connect", key="ConnectButton")
    if btn_connect:
        if acct_option == '<select account>':
            st.warning("Please select account")
        else:
            st.session_state.acc_con = snowflake_acct_connect(acct_option)  # get connection object without warehouse
            st.session_state.ws_con = snowflake_ws_connect(acct_option)  # get connection object without warehouse
            if st.session_state.acc_con is not None and st.session_state.ws_con is not None:
                st.success("Connected successfully")
            else:
                st.success("Connection failed")

    sql_text = st.text_area("Ad-Hoc SQL to execute on Snowflake")
    final_sql_list = sql_text.split(";") # Multiple sql statement separated by ; can be executed at once . else multiple SQl stateemrnt in a single API call is not supported

    btn_submit = st.button('Submit', key='AdhocSubmitButton')

    if btn_submit:
        with st.spinner("Executing SQL commands. Please wait..."):
            execute_sql(final_sql_list, "nxljbzs-qg26858", username, st.session_state.acc_con,st.session_state.ws_con)  # execute sql commandes
            st.success("Execution completed")

# Report section
def report():
    col1, col2 = st.columns([2, 1])
    # admin section to connect to an Account
    # with tab1:
    acct_list = get_ini_sections()  # get list of accounts from config.ini file
    acct_option = st.selectbox('Account', options=acct_list)
    btn_connect = st.button("Connect", "AccountButton")
    if btn_connect:
        if acct_option == '<select account>':
            st.warning("Please select account")
        else:
            st.session_state.acc_con = snowflake_acct_connect(acct_option)  # get connection object without warehouse
            st.session_state.ws_con = snowflake_ws_connect(acct_option)  # get connection object without warehouse
            if st.session_state.acc_con is not None and st.session_state.ws_con is not None:
                st.success("Connected successfully")
            else:
                st.success("Connection failed")
    if st.session_state.ws_con is not None:
        execdt_lst = get_execution_dates(st.session_state.ws_con)
        execdt_option = st.selectbox('Select date', options=execdt_lst)
        btn_report = st.button("Generate report", "GenerateReportButton")
        if btn_report:
            if st.session_state.ws_con is None:
                st.info("Please connect to an account")
            else:
                with st.spinner("Generating report. Please wait..."):
                    x = execdt_option.split("And")
                    y = x[0].split("=")
                    df_summary = generate_summ_report(str(y[1]).strip(), st.session_state.ws_con)
                    df_det = generate_det_report(str(y[1]).strip(), st.session_state.ws_con)
                    with col1:
                        st.write("***Summary***")
                        st.write("")
                        if df_summary.size > 0:
                            style = df_summary.style.hide_index()
                            st.write(style.to_html(), unsafe_allow_html=True)
                        st.write("")
                        st.write("***Detail***")
                        st.write("")
                        if df_det.size > 0:
                            style = df_det.style.hide_index()
                            st.write(style.to_html(), unsafe_allow_html=True)

                    with col2:
                        print("")
                        #lst_size = []
                        #labels = 'Success', 'Fail'
                        #if df_summary.size > 0:
                        #    for index, row in df_summary.iterrows():
                        #        lst_size.append(int(row[4]))
                        #        lst_size.append(int(row[5]))
                        #sizes = lst_size
                        #fig1, ax1 = plt.subplots()
                        #ax1.pie(sizes, labels=labels, autopct='%1.1f%%',
                        #        shadow=False, startangle=90)
                        #ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
                        #st.pyplot(fig1)
    else:
        st.info("Please connect to an account")
