import streamlit as st
from io import BytesIO
import pandas as pd
from generate_sql_from_template import generate_sql
import openpyxl as xl
from sql_execution import execute_sql, get_execution_dates, generate_summ_report, generate_det_report
from database_connection import snowflake_acct_connect, snowflake_ws_connect
from config import get_ini_sections
#pip install xlsxwriter
import numpy as np

def getSnowflakeObjectQueries():
    df_queries = pd.DataFrame()
    queryList = ["show warehouses",
                 "show databases",
                 "show schemas",
                 "show tables",
                 "show roles"]
    worksheetList = ["Warehouses",
                     "Databases",
                     "Schemas",
                     "Tables",
                     "Roles"]
    df_queries["sheets"] = worksheetList
    df_queries["queries"] = queryList
    return df_queries

def generateExcel(con):
    df_qry_list = getSnowflakeObjectQueries()
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    for index, row in df_qry_list.iterrows():
        sheetNm = row[0]
        qry = row[1]
        df = pd.read_sql(qry, con=con)

        df1 = df.copy()
        col_times = [col for col in df1.columns if any([isinstance(x, pd.Timestamp) for x in df1[col]])]
        for col in col_times:
            df1[col] = pd.to_datetime(df1[col], infer_datetime_format=True)
            df1[col] = df[col].dt.tz_localize(None)

        df1.to_excel(writer, index=False, sheet_name=sheetNm)

    writer.save()
    processed_data = output.getvalue()
    return processed_data

def DropSnowflakeObjects(username):
    uploaded_file = None
    # session objects to store snow flake connection
    if 'acc_con' not in st.session_state:
        st.session_state.acc_con = None

    if 'ws_con' not in st.session_state:
        st.session_state.ws_con = None
    acct_list = get_ini_sections()  # get list of accounts from config.ini file
    acct_option = st.selectbox('Account', options=acct_list)
    btn_connect = st.button("1.Connect", key='ConnectButton')
    if btn_connect:
        if acct_option == '<select account>':
            st.warning("Please select account")
        else:
            st.session_state.acc_con = snowflake_acct_connect(acct_option)  # get connection object without warehouse
            st.session_state.ws_con = snowflake_ws_connect(acct_option)  # get connection object without warehouse
            if st.session_state.acc_con is not None and st.session_state.ws_con is not None:
                st.success("Connected successfully")
                download_btn = st.download_button (label="2.Download snowflake objects",
                                                  data=generateExcel(st.session_state.ws_con),
                                                  file_name='snowflakeobjects.xlsx',
                                                  mime="application/vnd.ms-excel",
                                                  key="DownloadSnowflakeObjectButton")
            else:
                st.success("Connection failed")

    uploaded_file = st.file_uploader("Upload template", type=['xlsx'])

    btn_submit = st.button('3.Delete Objects', key='DeleteObjectButton')
    if btn_submit:
        if uploaded_file is None:
            st.info("Please upload template")
        elif st.session_state.acc_con is None:
            st.info("Please connect to an account using Admin tab")
        else:
            wb = xl.load_workbook(uploaded_file)
            sheets = wb.sheetnames
            res = len(wb.sheetnames)
            final_sql_list = []
            for i in range(res):
                if sheets[i] == "Source" :
                    with st.spinner("Generating SQL commands. Please wait..."):
                        df = pd.read_excel(
                            uploaded_file,
                            sheets[i]
                        )
                        df.replace(np.NaN, '', inplace=True)
                        final_sql_list= df['SQL Script'].tolist()
                        #st.write(final_sql_list)
            with st.spinner("Executing SQL commands. Please wait..."):
                execute_sql(final_sql_list, "nxljbzs-qg26858", username, st.session_state.acc_con,st.session_state.ws_con)  # execute sql commandes
                st.success("Execution completed")

