import time

import streamlit as st
import pandas as pd
import numpy as np
import requests
import snowflake.connector
import openpyxl as xl

from snowflake.connector import ProgrammingError

sql_cnt = 0
success_cnt = 0
fail_cnt = 0
all_sheets = []
uploaded_file = None

con = None
cur = None
df_status_report = pd.DataFrame()
if 'count' not in st.session_state:
    st.session_state.count = 0


def process():
    sql_list = generate_sql()
    execute_sql(sql_list)


def get_data_from_excel(filename, datasheet):
    df = pd.read_excel(
        filename,
        datasheet
    )
    df.replace(np.NaN, '', inplace=True)
    return df


def generate_sql():
    sql_list = []
    if uploaded_file is not None:

        df1 = get_data_from_excel(uploaded_file, 'schema')

        col_num = len(df1.columns)
        for index, row in df1.iterrows():
            sql_template = 'CREATE SCHEMA '
            for x in range(col_num):
                if row[x] != '':
                    if x == 0:
                        sql_template = sql_template + str(row[x]) + ' '
                    else:
                        if x == 1 and str(row[x] == 'Y'):
                            sql_template = sql_template + 'WITH MANAGED ACCESS '
                        else:
                            sql_template = sql_template + str(row.index[x]) + '=' + str(int(row[x])) + ' '
        df = pd.read_excel('c:\Temp\ObjectCreationTemplateV1.xlsx', sheet_name='Warehouse')
        df = df.replace(np.NaN, 'NA')
        # df=df1.isnull()
        # print(df)
        warehouse_template_mandatory_str = "CREATE OR REPLACE WAREHOUSE <I_WAREHOUSE_NAME>"
        warehouse_template_optional_str = "WITH WAREHOUSE_TYPE=<I_WAREHOUSE_TYPE> WAREHOUSE_SIZE=	<I_WAREHOUSE_SIZE> MAX_CLUSTER_COUNT =<I_MAX_CLUSTER_COUNT> MIN_CLUSTER_COUNT =<I_MIN_CLUSTER_COUNT> SCALING_POLICY =<I_SCALING_POLICY> AUTO_SUSPEND =<I_AUTO_SUSPEND> AUTO_RESUME=<I_AUTO_RESUME> INITIALLY_SUSPENDED=<I_INITIALLY_SUSPENDED> RESOURCE_MONITOR=<I_RESOURCE_MONITOR> COMMENT= <I_COMMENT> ENABLE_QUERY_ACCELERATION=<I_ENABLE_QUERY_ACCELERATION> QUERY_ACCELERATION_MAX_SCALE_FACTOR=<I_QUERY_ACCELERATION_MAX_SCALE_FACTOR> MAX_CONCURRENCY_LEVEL=<I_MAX_CONCURRENCY_LEVEL> STATEMENT_QUEUED_TIMEOUT_IN_SECONDS= <I_STATEMENT_QUEUED_TIMEOUT_IN_SECONDS> STATEMENT_TIMEOUT_IN_SECONDS=<I_STATEMENT_TIMEOUT_IN_SECONDS> TAG_NAME=<I_TAG_NAME> TAG_VALUE=<I_TAG_VALUE>"
        # print(warehouse_template_mandatory_str)
        # print(warehouse_template_optional_str)
        for index, row in df.iterrows():
            WAREHOUSE_NAME = row['WAREHOUSE_NAME']
            row_nonmandatory = row[1:]
            WAREHOUSE_TYPE = row['WAREHOUSE_TYPE']
            WAREHOUSE_SIZE = row['WAREHOUSE_SIZE']
            MAX_CLUSTER_COUNT = str(row['MAX_CLUSTER_COUNT'])
            MIN_CLUSTER_COUNT = str(row['MIN_CLUSTER_COUNT'])
            SCALING_POLICY = row['SCALING_POLICY']
            AUTO_SUSPEND = row['AUTO_SUSPEND']
            AUTO_RESUME = row['AUTO_RESUME']
            INITIALLY_SUSPENDED = row['INITIALLY_SUSPENDED']
            RESOURCE_MONITOR = row['RESOURCE_MONITOR']
            COMMENT = row['COMMENT']
            ENABLE_QUERY_ACCELERATION = row['ENABLE_QUERY_ACCELERATION']
            QUERY_ACCELERATION_MAX_SCALE_FACTOR = row['QUERY_ACCELERATION_MAX_SCALE_FACTOR']
            MAX_CONCURRENCY_LEVEL = row['MAX_CONCURRENCY_LEVEL']
            STATEMENT_QUEUED_TIMEOUT_IN_SECONDS = row['STATEMENT_QUEUED_TIMEOUT_IN_SECONDS']
            STATEMENT_TIMEOUT_IN_SECONDS = row['STATEMENT_TIMEOUT_IN_SECONDS']
            TAG_NAME = row['TAG_NAME']
            TAG_VALUE = row['TAG_VALUE']
            if (
                    WAREHOUSE_TYPE == 'NA' and WAREHOUSE_SIZE == 'NA' and MAX_CLUSTER_COUNT == 'NA' and MIN_CLUSTER_COUNT == 'NA' and SCALING_POLICY == 'NA' and AUTO_SUSPEND == 'NA' and AUTO_RESUME == 'NA' and INITIALLY_SUSPENDED == 'NA' and RESOURCE_MONITOR == 'NA' and COMMENT == 'NA' and ENABLE_QUERY_ACCELERATION == 'NA' and QUERY_ACCELERATION_MAX_SCALE_FACTOR == 'NA' and MAX_CONCURRENCY_LEVEL == 'NA' and STATEMENT_QUEUED_TIMEOUT_IN_SECONDS == 'NA' and STATEMENT_QUEUED_TIMEOUT_IN_SECONDS == 'NA' and TAG_NAME == 'NA' and TAG_VALUE == 'NA'):
                # if (math.isnan(WAREHOUSE_TYPE) and math.isnan(WAREHOUSE_SIZE) and math.isnan(MAX_CLUSTER_COUNT) and math.isnan(MIN_CLUSTER_COUNT) and math.isnan(SCALING_POLICY) and math.isnan(AUTO_SUSPEND) and math.isnan(AUTO_RESUME) and math.isnan(INITIALLY_SUSPENDED) and math.isnan(RESOURCE_MONITOR) and math.isnan(COMMENT) and math.isnan(ENABLE_QUERY_ACCELERATION) and math.isnan(QUERY_ACCELERATION_MAX_SCALE_FACTOR) and math.isnan(MAX_CONCURRENCY_LEVEL) and math.isnan(STATEMENT_QUEUED_TIMEOUT_IN_SECONDS) and math.isnan(STATEMENT_QUEUED_TIMEOUT_IN_SECONDS) and math.isnan(TAG_NAME) and math.isnan(TAG_VALUE)):
                # if row_nonmandatory.isnull().any(axis=1):
                mandatory_sql_str = warehouse_template_mandatory_str.replace("<I_WAREHOUSE_NAME>",
                                                                             WAREHOUSE_NAME)
                optional_sql_str = ""
            else:
                mandatory_sql_str = warehouse_template_mandatory_str.replace("<I_WAREHOUSE_NAME>",
                                                                             WAREHOUSE_NAME)
                if (WAREHOUSE_TYPE == 'NA'):
                    optional_sql_str = warehouse_template_optional_str.replace(
                        "WAREHOUSE_TYPE=<I_WAREHOUSE_TYPE>", "")
                else:
                    optional_sql_str = warehouse_template_optional_str.replace("<I_WAREHOUSE_TYPE>",
                                                                               WAREHOUSE_TYPE)
                if (WAREHOUSE_SIZE == 'NA'):
                    optional_sql_str = optional_sql_str.replace("WAREHOUSE_SIZE=<I_WAREHOUSE_SIZE>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_WAREHOUSE_SIZE>", WAREHOUSE_SIZE)
                if (MAX_CLUSTER_COUNT == 'NA'):
                    optional_sql_str = optional_sql_str.replace("MAX_CLUSTER_COUNT =<I_MAX_CLUSTER_COUNT>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_MAX_CLUSTER_COUNT>", MAX_CLUSTER_COUNT)
                if (MIN_CLUSTER_COUNT == 'NA'):
                    optional_sql_str = optional_sql_str.replace("MIN_CLUSTER_COUNT =<I_MIN_CLUSTER_COUNT>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_MIN_CLUSTER_COUNT>", MIN_CLUSTER_COUNT)
                if (SCALING_POLICY == 'NA'):
                    optional_sql_str = optional_sql_str.replace("SCALING_POLICY =<I_SCALING_POLICY>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_SCALING_POLICY>", SCALING_POLICY)
                if (AUTO_SUSPEND == 'NA'):
                    optional_sql_str = optional_sql_str.replace("AUTO_SUSPEND =<I_AUTO_SUSPEND>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_AUTO_SUSPEND>", AUTO_SUSPEND)
                if (AUTO_RESUME == 'NA'):
                    optional_sql_str = optional_sql_str.replace("AUTO_RESUME=<I_AUTO_RESUME>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_AUTO_RESUME>", AUTO_RESUME)
                if (INITIALLY_SUSPENDED == 'NA'):
                    optional_sql_str = optional_sql_str.replace("INITIALLY_SUSPENDED=<I_INITIALLY_SUSPENDED>",
                                                                "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_INITIALLY_SUSPENDED>", INITIALLY_SUSPENDED)
                if (RESOURCE_MONITOR == 'NA'):
                    optional_sql_str = optional_sql_str.replace("RESOURCE_MONITOR=<I_RESOURCE_MONITOR>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_RESOURCE_MONITOR>", RESOURCE_MONITOR)
                if (COMMENT == 'NA'):
                    optional_sql_str = optional_sql_str.replace("COMMENT= <I_COMMENT>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_COMMENT>", COMMENT)
                if (ENABLE_QUERY_ACCELERATION == 'NA'):
                    optional_sql_str = optional_sql_str.replace(
                        "ENABLE_QUERY_ACCELERATION=<I_ENABLE_QUERY_ACCELERATION>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_ENABLE_QUERY_ACCELERATION>",
                                                                ENABLE_QUERY_ACCELERATION)
                if (QUERY_ACCELERATION_MAX_SCALE_FACTOR == 'NA'):
                    optional_sql_str = optional_sql_str.replace(
                        "QUERY_ACCELERATION_MAX_SCALE_FACTOR=<I_QUERY_ACCELERATION_MAX_SCALE_FACTOR>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_QUERY_ACCELERATION_MAX_SCALE_FACTOR>",
                                                                QUERY_ACCELERATION_MAX_SCALE_FACTOR)
                if (MAX_CONCURRENCY_LEVEL == 'NA'):
                    optional_sql_str = optional_sql_str.replace(
                        "MAX_CONCURRENCY_LEVEL=<I_MAX_CONCURRENCY_LEVEL>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_MAX_CONCURRENCY_LEVEL>",
                                                                MAX_CONCURRENCY_LEVEL)
                if (STATEMENT_QUEUED_TIMEOUT_IN_SECONDS == 'NA'):
                    optional_sql_str = optional_sql_str.replace(
                        "STATEMENT_QUEUED_TIMEOUT_IN_SECONDS= <I_STATEMENT_QUEUED_TIMEOUT_IN_SECONDS>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_STATEMENT_QUEUED_TIMEOUT_IN_SECONDS>",
                                                                STATEMENT_QUEUED_TIMEOUT_IN_SECONDS)
                if (STATEMENT_TIMEOUT_IN_SECONDS == 'NA'):
                    optional_sql_str = optional_sql_str.replace(
                        "STATEMENT_TIMEOUT_IN_SECONDS=<I_STATEMENT_TIMEOUT_IN_SECONDS>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_STATEMENT_TIMEOUT_IN_SECONDS>",
                                                                STATEMENT_TIMEOUT_IN_SECONDS)
                if (TAG_NAME == 'NA'):
                    optional_sql_str = optional_sql_str.replace("TAG_NAME=<I_TAG_NAME>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_TAG_NAME>", TAG_NAME)
                if (TAG_VALUE == 'NA'):
                    optional_sql_str = optional_sql_str.replace("TAG_VALUE=<I_TAG_VALUE>", "")
                else:
                    optional_sql_str = optional_sql_str.replace("<I_TAG_VALUE>", TAG_VALUE)
            sql_template = mandatory_sql_str + " " + optional_sql_str + ";"

                    # st.write(sql_template)
            sql_list.append(sql_template)
    return sql_list


def api_authorize():
    session = requests.Session()


def snowflake_connect():
    ctx = snowflake.connector.connect(
        user='snowgliders',
        password='Infy@123',
        account='nxljbzs-qg26858',
        warehouse='SNOWFLAKE_HACKATHON_WS',
        database='SNOWFLAKE_POC'
    )
    return ctx


def execute_sql(sql_list):
    length = len(sql_list)
    global sql_cnt
    global success_cnt
    global fail_cnt
    qry_list = []
    status_list = []
    error_list = []
    global cur, con

    if length > 0:
        con = snowflake_connect()
        cur = con.cursor()
        for sql in sql_list:
            try:

                sql_cnt += 1
                qry_list.append(sql)
                cur.execute(sql)
                query_id = cur.sfqid
                con.get_query_status_throw_if_error(query_id)
                status_list.append("success")
                error_list.append("")
                success_cnt += 1
            except ProgrammingError as e:
                status_list.append("fail")
                error_list.append(e.msg)
                fail_cnt += 1
                continue
            except Exception as e:
                cur.close()
                con.close()

    df_status_report["SQL"] = qry_list
    df_status_report["STATUS"] = status_list
    df_status_report["ERROR"] = error_list
    if cur is not None:
        cur.close()
    if con is not None:
        con.close()


st.set_page_config(layout="wide")
st.header("Snowflake object creation utility")
sidebar1 = st.sidebar.header('Snowflake objects')
col1,col2=sidebar1.columns([2,1])
with col1:
    st.write('Snowflake objects')
    chk_warehouse=st.checkbox('Warehouse',key=1)
    chk_database=st.checkbox('Database',key=2)
    chk_schema=st.checkbox('Schema',key=3)



def get_sheets():
    # global uploaded_file
    # global all_sheets
    if uploaded_file is not None:
        st.write("hello")

        # x1 = pd.ExcelFile(uploaded_file)
        # all_sheets = x1.sheet_names
        # st.write(x1.sheet_names)
    # return all_sheets


uploaded_file = st.file_uploader("Upload template", type=['xlsx'])
if uploaded_file:
    wb = xl.load_workbook(uploaded_file)
    sheets = wb.sheetnames
    res = len(wb.sheetnames)
    for i in range(res):
        st.write(sheets[i])

btn_submit = st.button('Submit')
if btn_submit:
    chk_list=[]
    if chk_warehouse:
        chk_list.append("warehouse")
    if chk_database:
        chk_list.append("database")
    if chk_schema:
        chk_list.append("schema")

    chk_cnt=len(chk_list)
    with st.spinner("Executing SQL commands. Please wait..."):
        process()



if df_status_report.size > 0:
    st.write("***Execution summary***")
    st.write("SQL count: " + str(sql_cnt) + "  " + "Success count: " + str(success_cnt) + "  " + "Fail count: " + str(
        fail_cnt))

expander1 = st.expander(label='Execution details', expanded=False)
with expander1:
    if df_status_report.size > 0:
        style = df_status_report.style.hide_index()
        st.write(style.to_html(), unsafe_allow_html=True)
